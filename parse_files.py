from openpyxl import load_workbook
from bs4 import BeautifulSoup
from io import BytesIO

files = [
    ('diccionario_datos.xlsx','TestFiles/diccionario_datos.xlsx'),
    ('arquitectura_rag.drawio','TestFiles/arquitectura_rag.drawio'),
]

for name,path in files:
    print('===', name)
    with open(path,'rb') as f:
        content = f.read()
    if name.endswith('.xlsx'):
        try:
            wb = load_workbook(filename=BytesIO(content), data_only=True)
        except Exception as e:
            print('xlsx load error', e)
            wb=None
        parts=[]
        if wb:
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            parts.append(str(cell))
        text='\n'.join(parts)
        print('len',len(text))
        print(text[:1000])
    else:
        try:
            # Parse as XML for drawio files to avoid XML-as-HTML warnings
            soup = BeautifulSoup(content, 'xml')
            parts=[]
            for tag in soup.find_all():
                if tag.has_attr('value'):
                    parts.append(tag.get('value'))
            parts.extend([t.strip() for t in soup.stripped_strings])
            seen=set(); dedup=[]
            for p in parts:
                if not p: continue
                if p in seen: continue
                seen.add(p); dedup.append(p)
            text='\n'.join(dedup)
            print('len',len(text))
            print(text[:1000])
        except Exception as e:
            print('drawio parse error', e)
            print(content[:1000])
