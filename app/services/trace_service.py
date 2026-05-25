import csv
import json
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.models.context_chunk import ContextChunk
from app.schemas.trace_schema import (
    DataDictionaryIngestRequest,
    DataDictionaryIngestResponse,
    FieldUsageMatch,
    FieldUsageTraceRequest,
    FieldUsageTraceResponse,
)
from app.services.embedding_service import EmbeddingService


class TraceService:
    @staticmethod
    def ingest_data_dictionary(db: Session, payload: DataDictionaryIngestRequest) -> DataDictionaryIngestResponse:
        extension = Path(payload.file_path).suffix.lower()
        if extension == ".xlsx":
            return TraceService._ingest_xlsx(db, payload)
        if extension == ".csv":
            return TraceService._ingest_csv(db, payload)
        if extension == ".json":
            return TraceService._ingest_json(db, payload)
        raise ValueError("Unsupported dictionary file type. Use .xlsx, .csv, or .json")

    @staticmethod
    def _ingest_xlsx(db: Session, payload: DataDictionaryIngestRequest) -> DataDictionaryIngestResponse:
        workbook = load_workbook(payload.file_path, read_only=True, data_only=True)
        target_sheets = [payload.sheet_name] if payload.sheet_name else workbook.sheetnames

        total_fields = 0
        processed_sheets = 0

        for sheet_name in target_sheets:
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]
            rows = ws.iter_rows(values_only=True)

            try:
                headers = [str(cell).strip().lower() if cell is not None else "" for cell in next(rows)]
            except StopIteration:
                continue

            field_idx = TraceService._find_header_index(headers, ["field", "campo", "column", "name"])
            if field_idx is None:
                continue

            definition_idx = TraceService._find_header_index(
                headers,
                ["definition", "descripcion", "description", "meaning", "detalle"],
            )

            chunks: list[ContextChunk] = []
            row_count = 1
            for row in rows:
                row_count += 1
                if row_count > payload.max_rows:
                    break

                field_value = TraceService._cell_to_text(row, field_idx)
                if not field_value:
                    continue

                definition_value = TraceService._cell_to_text(row, definition_idx) if definition_idx is not None else ""
                content = f"Field: {field_value}\nDefinition: {definition_value}".strip()
                embedding = EmbeddingService.embed_text(content)
                chunks.append(
                    ContextChunk(
                        source=payload.dictionary_name,
                        content=content,
                        metadata_json={
                            "document_type": "data_dictionary",
                            "dictionary_name": payload.dictionary_name,
                            "tab": sheet_name,
                            "row_number": row_count,
                            "field_name": field_value,
                            "definition": definition_value,
                        },
                        embedding=embedding,
                    )
                )

            if chunks:
                db.add_all(chunks)
                db.commit()
                total_fields += len(chunks)
                processed_sheets += 1

        return DataDictionaryIngestResponse(
            dictionary_name=payload.dictionary_name,
            sheets_processed=processed_sheets,
            fields_ingested=total_fields,
        )

    @staticmethod
    def _ingest_csv(db: Session, payload: DataDictionaryIngestRequest) -> DataDictionaryIngestResponse:
        with open(payload.file_path, "r", encoding="utf-8", errors="ignore", newline="") as csv_file:
            reader = csv.reader(csv_file)
            try:
                headers = [str(cell).strip().lower() if cell is not None else "" for cell in next(reader)]
            except StopIteration:
                return DataDictionaryIngestResponse(
                    dictionary_name=payload.dictionary_name,
                    sheets_processed=0,
                    fields_ingested=0,
                )

            field_idx = TraceService._find_header_index(headers, ["field", "campo", "column", "name"])
            if field_idx is None:
                raise ValueError("CSV dictionary requires a field/campo/column/name header")

            definition_idx = TraceService._find_header_index(
                headers,
                ["definition", "descripcion", "description", "meaning", "detalle"],
            )

            chunks: list[ContextChunk] = []
            for row_count, row in enumerate(reader, start=2):
                if row_count > payload.max_rows:
                    break

                field_value = TraceService._cell_to_text(row, field_idx)
                if not field_value:
                    continue

                definition_value = TraceService._cell_to_text(row, definition_idx) if definition_idx is not None else ""
                content = f"Field: {field_value}\nDefinition: {definition_value}".strip()
                embedding = EmbeddingService.embed_text(content)
                chunks.append(
                    ContextChunk(
                        source=payload.dictionary_name,
                        content=content,
                        metadata_json={
                            "document_type": "data_dictionary",
                            "dictionary_name": payload.dictionary_name,
                            "tab": payload.sheet_name or "csv",
                            "row_number": row_count,
                            "field_name": field_value,
                            "definition": definition_value,
                        },
                        embedding=embedding,
                    )
                )

            if chunks:
                db.add_all(chunks)
                db.commit()

        return DataDictionaryIngestResponse(
            dictionary_name=payload.dictionary_name,
            sheets_processed=1,
            fields_ingested=len(chunks),
        )

    @staticmethod
    def _ingest_json(db: Session, payload: DataDictionaryIngestRequest) -> DataDictionaryIngestResponse:
        with open(payload.file_path, "r", encoding="utf-8", errors="ignore") as json_file:
            parsed = json.load(json_file)

        records: list[dict] = []
        if isinstance(parsed, list):
            records = [record for record in parsed if isinstance(record, dict)]
        elif isinstance(parsed, dict):
            if isinstance(parsed.get("fields"), list):
                records = [record for record in parsed.get("fields", []) if isinstance(record, dict)]
            else:
                records = [parsed]

        chunks: list[ContextChunk] = []
        for row_count, record in enumerate(records, start=1):
            if row_count > payload.max_rows:
                break

            field_value = str(
                record.get("field")
                or record.get("campo")
                or record.get("name")
                or record.get("column")
                or ""
            ).strip()
            if not field_value:
                continue

            definition_value = str(
                record.get("definition")
                or record.get("descripcion")
                or record.get("description")
                or record.get("meaning")
                or ""
            ).strip()

            content = f"Field: {field_value}\nDefinition: {definition_value}".strip()
            embedding = EmbeddingService.embed_text(content)
            chunks.append(
                ContextChunk(
                    source=payload.dictionary_name,
                    content=content,
                    metadata_json={
                        "document_type": "data_dictionary",
                        "dictionary_name": payload.dictionary_name,
                        "tab": payload.sheet_name or "json",
                        "row_number": row_count,
                        "field_name": field_value,
                        "definition": definition_value,
                    },
                    embedding=embedding,
                )
            )

        if chunks:
            db.add_all(chunks)
            db.commit()

        return DataDictionaryIngestResponse(
            dictionary_name=payload.dictionary_name,
            sheets_processed=1,
            fields_ingested=len(chunks),
        )

    @staticmethod
    def trace_field_usage(db: Session, payload: FieldUsageTraceRequest) -> FieldUsageTraceResponse:
        dictionary_row = db.execute(
            text(
                """
                SELECT metadata_json->>'tab' AS tab,
                       CAST(metadata_json->>'row_number' AS INTEGER) AS row_number,
                       metadata_json->>'definition' AS definition
                FROM context_chunks
                WHERE metadata_json->>'document_type' = 'data_dictionary'
                  AND metadata_json->>'dictionary_name' = :dictionary_name
                  AND LOWER(metadata_json->>'field_name') = LOWER(:field_name)
                ORDER BY id DESC
                LIMIT 1
                """
            ),
            {"dictionary_name": payload.dictionary_name, "field_name": payload.field_name},
        ).mappings().first()

        if not dictionary_row:
            raise ValueError("Field not found in the specified data dictionary")

        rows = db.execute(
            text(
                """
                SELECT id,
                       source,
                       content,
                       metadata_json,
                       (embedding <=> CAST(:embedding AS vector)) AS distance
                FROM context_chunks
                WHERE metadata_json->>'document_type' = 'code'
                  AND content ILIKE :pattern
                ORDER BY distance ASC
                LIMIT :limit
                """
            ),
            {
                "pattern": f"%{payload.field_name}%",
                "limit": payload.top_k,
                "embedding": str(EmbeddingService.embed_text(payload.field_name)),
            },
        ).mappings().all()

        matches: list[FieldUsageMatch] = []
        for row in rows:
            metadata = row["metadata_json"] or {}
            file_path = metadata.get("file_path")
            if not file_path:
                continue
            similarity = max(0.0, 1.0 - float(row["distance"]))
            matches.append(
                FieldUsageMatch(
                    source=row["source"],
                    file_path=file_path,
                    line_start=int(metadata.get("line_start", 1)),
                    line_end=int(metadata.get("line_end", 1)),
                    similarity=similarity,
                    snippet=str(row["content"])[:400],
                )
            )

        return FieldUsageTraceResponse(
            field_name=payload.field_name,
            dictionary_name=payload.dictionary_name,
            dictionary_tab=dictionary_row["tab"],
            dictionary_row=dictionary_row["row_number"],
            dictionary_definition=dictionary_row["definition"],
            matches=matches,
        )

    @staticmethod
    def _find_header_index(headers: list[str], candidates: list[str]) -> int | None:
        for idx, header in enumerate(headers):
            for candidate in candidates:
                if candidate in header:
                    return idx
        return None

    @staticmethod
    def _cell_to_text(row, idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        value = row[idx]
        if value is None:
            return ""
        return str(value).strip()
