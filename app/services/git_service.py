import csv
import json
from pathlib import Path
import re
import shutil
import tempfile
from urllib.parse import urlparse

from git import Repo
from git.exc import GitCommandError, InvalidGitRepositoryError, NoSuchPathError
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.orm import Session

from app.models.context_chunk import ContextChunk
from app.schemas.git_schema import (
    GitCommitSummary,
    GitIngestRequest,
    GitIngestResponse,
    GitScanResponse,
    GitSourceRequest,
)
from app.services.embedding_service import EmbeddingService


class GitService:
    @staticmethod
    def scan_repository(payload: GitSourceRequest) -> GitScanResponse:
        temp_dir: str | None = None

        try:
            repo, repo_root, source_type, repo_name, temp_dir = GitService._prepare_repository(payload)

            extensions = GitService._normalize_extensions(payload.include_extensions)
            files = GitService._collect_files(repo_root, extensions, payload.max_files)

            head_commit = repo.head.commit
            branch_name = repo.git.rev_parse("--abbrev-ref", "HEAD")

            return GitScanResponse(
                repo_name=repo_name,
                source_type=source_type,
                current_branch=branch_name,
                latest_commit=GitCommitSummary(
                    hash=head_commit.hexsha,
                    author=str(head_commit.author),
                    date=head_commit.committed_datetime.isoformat(),
                    message=head_commit.message.strip(),
                ),
                scanned_files_total=len(files),
                sampled_files=files,
            )
        except (InvalidGitRepositoryError, NoSuchPathError) as exc:
            raise ValueError("The provided local path is not a valid Git repository") from exc
        except GitCommandError as exc:
            raise RuntimeError(f"Unable to access Git repository: {exc}") from exc
        finally:
            if temp_dir:
                shutil.rmtree(temp_dir, ignore_errors=True)

    @staticmethod
    def ingest_repository(payload: GitIngestRequest, db: Session) -> GitIngestResponse:
        temp_dir: str | None = None

        try:
            repo, repo_root, source_type, repo_name, temp_dir = GitService._prepare_repository(payload)

            extensions = GitService._normalize_extensions(payload.include_extensions)
            files = GitService._collect_files(repo_root, extensions, payload.max_files)

            head_commit = repo.head.commit
            branch_name = repo.git.rev_parse("--abbrev-ref", "HEAD")

            chunks_to_insert: list[ContextChunk] = []
            files_processed = 0

            for rel_path in files:
                full_path = repo_root / rel_path
                file_chunks = GitService._extract_chunks_for_file(
                    full_path=full_path,
                    rel_path=rel_path,
                    chunk_size=payload.chunk_size,
                    chunk_overlap=payload.chunk_overlap,
                    max_chunks=payload.max_chunks_per_file,
                )
                if not file_chunks:
                    continue

                files_processed += 1
                for chunk_index, chunk_info in enumerate(file_chunks):
                    chunk = chunk_info["content"]
                    embedding = EmbeddingService.embed_text(chunk)
                    chunks_to_insert.append(
                        ContextChunk(
                            source=repo_name,
                            content=chunk,
                            metadata_json={
                                "document_type": chunk_info["document_type"],
                                "artifact_type": chunk_info["artifact_type"],
                                "source_type": source_type,
                                "file_path": rel_path,
                                "chunk_index": chunk_index,
                                "line_start": chunk_info["line_start"],
                                "line_end": chunk_info["line_end"],
                                "tab": chunk_info.get("tab"),
                                "page": chunk_info.get("page"),
                                "branch": branch_name,
                                "commit": head_commit.hexsha,
                            },
                            embedding=embedding,
                        )
                    )

            if chunks_to_insert:
                db.add_all(chunks_to_insert)
                db.commit()

            return GitIngestResponse(
                repo_name=repo_name,
                source_type=source_type,
                current_branch=branch_name,
                latest_commit=GitCommitSummary(
                    hash=head_commit.hexsha,
                    author=str(head_commit.author),
                    date=head_commit.committed_datetime.isoformat(),
                    message=head_commit.message.strip(),
                ),
                files_processed=files_processed,
                chunks_inserted=len(chunks_to_insert),
            )
        except (InvalidGitRepositoryError, NoSuchPathError) as exc:
            raise ValueError("The provided local path is not a valid Git repository") from exc
        except GitCommandError as exc:
            raise RuntimeError(f"Unable to access Git repository: {exc}") from exc
        finally:
            if temp_dir:
                shutil.rmtree(temp_dir, ignore_errors=True)

    @staticmethod
    def _prepare_repository(payload: GitSourceRequest) -> tuple[Repo, Path, str, str, str | None]:
        temp_dir: str | None = None
        if payload.repo_url:
            temp_dir = tempfile.mkdtemp(prefix="rag_git_")
            clone_kwargs: dict[str, str | int | bool] = {"depth": 1, "single_branch": True}
            if payload.branch:
                clone_kwargs["branch"] = payload.branch
            repo = Repo.clone_from(payload.repo_url, temp_dir, **clone_kwargs)
            repo_root = Path(temp_dir)
            source_type = "remote"
            repo_name = GitService._repo_name_from_url(payload.repo_url)
            return repo, repo_root, source_type, repo_name, temp_dir

        local_path = Path(payload.local_path or "")
        if not local_path.exists():
            raise ValueError(f"Local path does not exist: {local_path}")

        repo = Repo(local_path)
        repo_root = Path(repo.working_tree_dir or str(local_path))
        source_type = "local"
        repo_name = repo_root.name
        return repo, repo_root, source_type, repo_name, temp_dir

    @staticmethod
    def _normalize_extensions(extensions: list[str]) -> set[str]:
        normalized: set[str] = set()
        for ext in extensions:
            clean_ext = ext.strip().lower()
            if not clean_ext:
                continue
            if not clean_ext.startswith("."):
                clean_ext = f".{clean_ext}"
            normalized.add(clean_ext)
        return normalized

    @staticmethod
    def _collect_files(repo_root: Path, extensions: set[str], limit: int) -> list[str]:
        results: list[str] = []
        special_filenames = {"dockerfile", "docker-compose.yml", "docker-compose.yaml"}

        for file_path in repo_root.rglob("*"):
            if len(results) >= limit:
                break
            if not file_path.is_file():
                continue
            if ".git" in file_path.parts:
                continue
            file_name = file_path.name.lower()
            if file_path.suffix.lower() not in extensions and file_name not in special_filenames:
                continue

            rel_path = file_path.relative_to(repo_root).as_posix()
            results.append(rel_path)

        return results

    @staticmethod
    def _repo_name_from_url(repo_url: str) -> str:
        parsed = urlparse(repo_url)
        name = parsed.path.rstrip("/").split("/")[-1]
        if name.endswith(".git"):
            name = name[:-4]
        return name or "repository"

    @staticmethod
    def _read_text_file(file_path: Path) -> str:
        try:
            return file_path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            return ""

    @staticmethod
    def _extract_chunks_for_file(
        full_path: Path,
        rel_path: str,
        chunk_size: int,
        chunk_overlap: int,
        max_chunks: int,
    ) -> list[dict[str, int | str | None]]:
        suffix = full_path.suffix.lower()
        file_name = full_path.name.lower()

        if suffix in {".xlsx"}:
            return GitService._extract_xlsx_chunks(full_path, max_chunks)
        if suffix in {".csv"}:
            return GitService._extract_csv_chunks(full_path, max_chunks)
        if suffix in {".json"}:
            return GitService._extract_json_chunks(full_path, chunk_size, chunk_overlap, max_chunks)
        if suffix in {".pdf"}:
            return GitService._extract_pdf_chunks(full_path, chunk_size, chunk_overlap, max_chunks)
        if suffix in {".drawio", ".xml"} and "drawio" in rel_path.lower():
            return GitService._extract_drawio_chunks(full_path, chunk_size, chunk_overlap, max_chunks)

        content = GitService._read_text_file(full_path)
        if not content:
            return []

        artifact_type = GitService._classify_artifact_type(suffix, file_name, rel_path, content)
        document_type = "code" if artifact_type == "code" else "artifact"

        chunks = GitService._split_text_with_lines(content, chunk_size, chunk_overlap, max_chunks)
        for chunk in chunks:
            chunk["document_type"] = document_type
            chunk["artifact_type"] = artifact_type
            chunk["tab"] = None
            chunk["page"] = None
        return chunks

    @staticmethod
    def _classify_artifact_type(suffix: str, file_name: str, rel_path: str, content: str) -> str:
        path_lower = rel_path.lower()
        if suffix in {".drawio", ".mmd", ".mermaid"} or "drawio" in path_lower:
            return "architecture"
        if file_name in {"dockerfile", "docker-compose.yml", "docker-compose.yaml"}:
            return "infrastructure"
        if suffix in {".yaml", ".yml"}:
            if any(token in path_lower for token in ["k8s", "kubernetes", "helm", "manifests"]):
                return "infrastructure"
            if any(token in content.lower() for token in ["apiVersion:", "kind:", "metadata:"]):
                return "infrastructure"
            return "configuration"
        if suffix in {".xlsx", ".csv"}:
            return "data_dictionary"
        if suffix == ".json":
            if any(token in path_lower for token in ["dictionary", "diccionario", "schema", "fields"]):
                return "data_dictionary"
            return "documentation"
        if suffix in {".md", ".pdf"}:
            return "documentation"
        if suffix in {".py", ".js", ".ts", ".tsx", ".jsx", ".java", ".cs", ".go", ".rs", ".php", ".rb"}:
            return "code"
        return "artifact"

    @staticmethod
    def _extract_xlsx_chunks(full_path: Path, max_chunks: int) -> list[dict[str, int | str | None]]:
        workbook = load_workbook(full_path, read_only=True, data_only=True)
        chunks: list[dict[str, int | str | None]] = []

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if len(chunks) >= max_chunks:
                    return chunks
                values = ["" if cell is None else str(cell).strip() for cell in row]
                text = " | ".join([v for v in values if v])
                if not text:
                    continue
                chunks.append(
                    {
                        "content": text,
                        "line_start": row_idx,
                        "line_end": row_idx,
                        "document_type": "artifact",
                        "artifact_type": "data_dictionary",
                        "tab": sheet_name,
                        "page": None,
                    }
                )
        return chunks

    @staticmethod
    def _extract_csv_chunks(full_path: Path, max_chunks: int) -> list[dict[str, int | str | None]]:
        chunks: list[dict[str, int | str | None]] = []
        try:
            with full_path.open("r", encoding="utf-8", errors="ignore", newline="") as csv_file:
                reader = csv.reader(csv_file)
                for row_idx, row in enumerate(reader, start=1):
                    if len(chunks) >= max_chunks:
                        break
                    text = " | ".join([cell.strip() for cell in row if cell and cell.strip()])
                    if not text:
                        continue
                    chunks.append(
                        {
                            "content": text,
                            "line_start": row_idx,
                            "line_end": row_idx,
                            "document_type": "artifact",
                            "artifact_type": "data_dictionary",
                            "tab": None,
                            "page": None,
                        }
                    )
        except OSError:
            return []
        return chunks

    @staticmethod
    def _extract_json_chunks(
        full_path: Path,
        chunk_size: int,
        chunk_overlap: int,
        max_chunks: int,
    ) -> list[dict[str, int | str | None]]:
        try:
            raw = full_path.read_text(encoding="utf-8", errors="ignore")
            parsed = json.loads(raw)
            normalized = json.dumps(parsed, ensure_ascii=False, indent=2)
        except (OSError, json.JSONDecodeError):
            return []

        chunks = GitService._split_text_with_lines(normalized, chunk_size, chunk_overlap, max_chunks)
        for chunk in chunks:
            chunk["document_type"] = "artifact"
            chunk["artifact_type"] = "data_dictionary"
            chunk["tab"] = None
            chunk["page"] = None
        return chunks

    @staticmethod
    def _extract_pdf_chunks(
        full_path: Path,
        chunk_size: int,
        chunk_overlap: int,
        max_chunks: int,
    ) -> list[dict[str, int | str | None]]:
        try:
            reader = PdfReader(str(full_path))
        except Exception:
            return []

        chunks: list[dict[str, int | str | None]] = []
        for page_idx, page in enumerate(reader.pages, start=1):
            if len(chunks) >= max_chunks:
                break
            text = page.extract_text() or ""
            if not text.strip():
                continue
            page_chunks = GitService._split_text_with_lines(text, chunk_size, chunk_overlap, max_chunks - len(chunks))
            for chunk in page_chunks:
                chunk["document_type"] = "artifact"
                chunk["artifact_type"] = "documentation"
                chunk["tab"] = None
                chunk["page"] = page_idx
                chunks.append(chunk)
                if len(chunks) >= max_chunks:
                    break
        return chunks

    @staticmethod
    def _extract_drawio_chunks(
        full_path: Path,
        chunk_size: int,
        chunk_overlap: int,
        max_chunks: int,
    ) -> list[dict[str, int | str | None]]:
        raw = GitService._read_text_file(full_path)
        if not raw:
            return []

        labels = re.findall(r'value="([^"]+)"', raw)
        text_content = "\n".join(labels) if labels else raw
        chunks = GitService._split_text_with_lines(text_content, chunk_size, chunk_overlap, max_chunks)
        for chunk in chunks:
            chunk["document_type"] = "artifact"
            chunk["artifact_type"] = "architecture"
            chunk["tab"] = None
            chunk["page"] = None
        return chunks

    @staticmethod
    def _split_text_with_lines(
        text: str,
        chunk_size: int,
        chunk_overlap: int,
        max_chunks: int,
    ) -> list[dict[str, int | str]]:
        lines = text.splitlines()
        if not lines:
            return []

        results: list[dict[str, int | str]] = []
        start_idx = 0
        overlap_lines = 0 if chunk_overlap == 0 else max(1, chunk_overlap // 120)

        while start_idx < len(lines) and len(results) < max_chunks:
            end_idx = start_idx
            current_size = 0

            while end_idx < len(lines):
                line = lines[end_idx]
                current_size += len(line) + 1
                end_idx += 1
                if current_size >= chunk_size:
                    break

            chunk_lines = lines[start_idx:end_idx]
            chunk_text = "\n".join(chunk_lines).strip()
            if chunk_text:
                results.append(
                    {
                        "content": chunk_text,
                        "line_start": start_idx + 1,
                        "line_end": end_idx,
                    }
                )

            if end_idx >= len(lines):
                break
            start_idx = max(start_idx + 1, end_idx - overlap_lines)

        return results
